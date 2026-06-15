import glob
import json
import os
import re
import smtplib
import ssl
import sys
import time
from datetime import datetime
from email.header import Header
from email.mime.text import MIMEText

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT_DIR, ".env"))
if str(os.getenv("GITHUB_ACTIONS") or "").strip().lower() not in {"1", "true", "yes", "y", "on"}:
    load_dotenv(os.path.join(ROOT_DIR, "h4", ".env"), override=True)

for stream_name in ("stdout", "stderr"):
    stream = getattr(sys, stream_name, None)
    if hasattr(stream, "reconfigure"):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


os.environ.setdefault("TZ", "Asia/Shanghai")
try:
    time.tzset()
except Exception:
    pass

RISK_CONTROL_MESSAGE = os.getenv("RISK_CONTROL_MESSAGE", "抽奖失败，疑似触发活动限制").strip()

STATUS_RED_FILL = PatternFill("solid", fgColor="F8696B")
STATUS_BLUE_FILL = PatternFill("solid", fgColor="9DC3E6")
FONT_GREEN = Font(color="008000")
FONT_RED = Font(color="C00000")
FONT_DARK = Font(color="000000")
BAD_LOTTERY_TITLE_RE = re.compile(r"(抽奖机会|我的抽奖机会|立即抽奖|开始抽奖|去抽奖|再抽一次|报名|兑换|活动规则|订单统计)")


def truthy(value) -> bool:
    if value is True:
        return True
    if value is False or value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def safe_int(value, default=0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def safe_float(value, default=0.0) -> float:
    try:
        return float(str(value).strip())
    except Exception:
        return default


def pick_money_value(*values) -> float:
    fallback = None
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        money = safe_float(text, 0.0)
        if money != 0:
            return money
        if fallback is None:
            fallback = 0.0
    return fallback if fallback is not None else 0.0


def record_invoice_money(record: dict) -> float:
    return pick_money_value(
        record.get("invoice_money"),
        record.get("consumption_amount"),
        record.get("invoiceMoney"),
    )


def default_group_name(group_number: int) -> str:
    return f"{group_number}组" if group_number > 0 else ""


def default_group_position(group_number: int, account_index: int) -> str:
    if group_number > 0 and account_index > 0:
        return f"{group_number}组账号{account_index}"
    return f"账号{account_index}" if account_index > 0 else "未知账号"


def load_account_lookup() -> tuple[dict[tuple[int, int], str], int]:
    lookup = {}
    total = 0
    for group_number in range(1, 5):
        raw = os.getenv(f"ACCOUNTS_BATCH{group_number}", "") or ""
        for account_index, line in enumerate(raw.splitlines(), start=1):
            line = line.strip()
            if not line or "," not in line:
                continue
            username = line.split(",", 1)[0].strip()
            lookup[(group_number, account_index)] = username
            total += 1
    return lookup, total


def load_manifest(results_dir: str) -> dict:
    path = os.path.join(results_dir, "manifest.json")
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return {}


def clean_lottery_title(title: str) -> str:
    text = re.sub(r"\s+", " ", str(title or "")).strip()
    compacted = re.sub(r"\s+", "", text)
    if not text or BAD_LOTTERY_TITLE_RE.search(compacted):
        return ""
    return text


def target_date_text(manifest: dict) -> str:
    if isinstance(manifest, dict) and manifest.get("target_date"):
        return str(manifest["target_date"]).strip()
    return datetime.now().strftime("%Y-%m-%d")


def resolve_output_xlsx_path(results_dir: str, manifest: dict) -> str:
    filename = f"{target_date_text(manifest)}抽奖汇总.xlsx"
    configured_path = (os.getenv("OUTPUT_XLSX_PATH") or "").strip()
    if configured_path:
        return configured_path
    return os.path.join(results_dir, filename)


def find_json_files(results_dir: str) -> list[str]:
    paths = set()
    for path in glob.glob(os.path.join(results_dir, "**", "*.json"), recursive=True):
        if os.path.basename(path).lower() == "manifest.json":
            continue
        if os.path.isfile(path):
            paths.add(path)
    return sorted(paths)


def record_key(record: dict):
    group_number = safe_int(record.get("group_number"), 0)
    account_index = safe_int(record.get("account_index"), 0)
    if group_number > 0 and account_index > 0:
        return group_number, account_index
    return None


def normalize_activity_records(value) -> dict:
    if not isinstance(value, dict):
        return {"lottery": []}

    normalized = {"lottery": []}
    rows = value.get("lottery")
    if not isinstance(rows, list):
        rows = []
    for item in rows:
        if len(normalized["lottery"]) >= 3:
            break
        if not isinstance(item, dict):
            continue
        title = clean_lottery_title(
            item.get("title")
            or item.get("prize_title")
            or item.get("prize_name")
            or item.get("skuTitle")
            or item.get("prizeTitle")
            or item.get("goodsName")
            or item.get("awardName")
        )
        if not title:
            continue
        normalized["lottery"].append(
            {
                "title": title,
                "status_text": str(item.get("status_text") or "").strip(),
                "claimed": truthy(item.get("claimed")),
                "expiry_date": str(item.get("expiry_date") or "").strip(),
                "won_at": str(item.get("won_at") or "").strip(),
            }
        )
    return normalized


def normalize_record(record: dict, payload: dict, account_lookup: dict[tuple[int, int], str]) -> dict:
    group_number = safe_int(record.get("group_number", payload.get("group_number")), 0)
    account_index = safe_int(record.get("account_index"), 0)
    username = str(
        record.get("username")
        or record.get("masked_username")
        or account_lookup.get((group_number, account_index))
        or f"账号{account_index}"
    ).strip()
    detail_reason = str(record.get("detail_reason") or "").strip()
    risk_controlled = truthy(record.get("risk_controlled")) or (RISK_CONTROL_MESSAGE and RISK_CONTROL_MESSAGE in detail_reason)
    group_name = str(record.get("group_name") or payload.get("group_name") or payload.get("batch_name") or default_group_name(group_number)).strip()
    group_position = str(record.get("group_position") or default_group_position(group_number, account_index)).strip()
    return {
        "account_index": account_index,
        "execution_order": safe_int(record.get("execution_order"), 0),
        "username": username,
        "group_name": group_name,
        "group_number": group_number,
        "group_position": group_position,
        "sign_success": truthy(record.get("sign_success")),
        "sign_status": str(record.get("sign_status") or "").strip(),
        "initial_points": safe_float(record.get("initial_points"), 0.0),
        "final_points": safe_float(record.get("final_points"), 0.0),
        "points_reward": safe_float(record.get("points_reward"), 0.0),
        "invoice_money": record_invoice_money(record),
        "has_reward": truthy(record.get("has_reward")),
        "password_error": truthy(record.get("password_error")),
        "risk_controlled": risk_controlled,
        "retry_count": safe_int(record.get("retry_count"), 0),
        "is_final_retry": truthy(record.get("is_final_retry")),
        "detail_reason": detail_reason,
        "sign_time": str(record.get("sign_time") or "").strip(),
        "sign_ip": str(record.get("sign_ip") or "").strip(),
        "activity_records": normalize_activity_records(record.get("activity_records")),
    }


def load_results(results_dir: str, account_lookup: dict[tuple[int, int], str]) -> list[dict]:
    records_by_key = {}
    extras = []
    for path in find_json_files(results_dir):
        try:
            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)
        except Exception:
            continue
        rows = payload.get("results") if isinstance(payload, dict) else None
        if not isinstance(rows, list):
            continue
        for record in rows:
            if not isinstance(record, dict):
                continue
            normalized = normalize_record(record, payload, account_lookup)
            key = record_key(normalized)
            if key is None:
                extras.append(normalized)
            else:
                records_by_key[key] = normalized
    return list(records_by_key.values()) + extras


def build_missing_record(group_number: int, account_index: int, username: str) -> dict:
    return {
        "account_index": account_index,
        "execution_order": account_index,
        "username": username,
        "group_name": default_group_name(group_number),
        "group_number": group_number,
        "group_position": default_group_position(group_number, account_index),
        "sign_success": False,
        "sign_status": "抽奖异常",
        "initial_points": 0.0,
        "final_points": 0.0,
        "points_reward": 0.0,
        "invoice_money": 0.0,
        "has_reward": False,
        "password_error": False,
        "risk_controlled": False,
        "retry_count": 0,
        "is_final_retry": False,
        "detail_reason": "缺少抽奖结果",
        "sign_time": "",
        "sign_ip": "",
        "activity_records": {"lottery": []},
    }


def merge_records_with_expected(records: list[dict], account_lookup: dict[tuple[int, int], str]) -> list[dict]:
    indexed = {}
    extras = []
    for record in records:
        key = record_key(record)
        if key is None:
            extras.append(record)
        else:
            indexed[key] = record

    if not account_lookup:
        return list(indexed.values()) + extras

    merged = []
    for key in sorted(account_lookup):
        record = indexed.pop(key, None)
        if record is None:
            merged.append(build_missing_record(key[0], key[1], account_lookup[key]))
            continue
        if not record.get("username"):
            record["username"] = account_lookup[key]
        merged.append(record)

    unexpected = sorted(
        list(indexed.values()) + extras,
        key=lambda item: (
            safe_int(item.get("group_number"), 999999),
            safe_int(item.get("account_index"), 999999),
            str(item.get("username") or ""),
        ),
    )
    merged.extend(unexpected)
    return merged


def status_label(record: dict) -> str:
    raw_status = str(record.get("sign_status") or "")
    if truthy(record.get("risk_controlled")):
        return "抽奖风控"
    if truthy(record.get("sign_success")):
        return "抽奖成功"
    if truthy(record.get("password_error")) or any(keyword in raw_status for keyword in ("失败", "错误", "Token", "token")):
        return "抽奖失败"
    return "抽奖异常"


def detail_reason(record: dict) -> str:
    reason = str(record.get("detail_reason") or "").strip()
    if reason:
        return reason
    if record.get("risk_controlled"):
        return RISK_CONTROL_MESSAGE
    if record.get("sign_status"):
        return str(record["sign_status"]).strip()
    return "抽奖异常"


def detail_text(record: dict) -> str:
    if truthy(record.get("sign_success")):
        return str(record.get("sign_status") or "抽奖成功").strip()
    return detail_reason(record)


def is_problem_record(record: dict) -> bool:
    return status_sort_bucket(record) == 0


def status_sort_bucket(record: dict) -> int:
    label = status_label(record)
    if label in {"抽奖失败", "抽奖异常", "抽奖风控"}:
        return 0
    return 2


def sort_records(records: list[dict]) -> list[dict]:
    return sorted(
        records,
        key=lambda item: (
            0 if is_problem_record(item) else 1,
            -lottery_count(item),
            safe_int(item.get("group_number"), 999999),
            safe_int(item.get("account_index"), 999999),
            str(item.get("username") or ""),
        ),
    )


def format_percent(value: float) -> str:
    return f"{value:.1f}".rstrip("0").rstrip(".")


def lottery_items(record: dict) -> list[dict]:
    activity = normalize_activity_records(record.get("activity_records"))
    return [item for item in (activity.get("lottery") or []) if str(item.get("title") or "").strip()]


def lottery_count(record: dict) -> int:
    return len(lottery_items(record))


def prize_distribution(records: list[dict]) -> dict[str, int]:
    distribution = {}
    for record in records:
        for item in lottery_items(record):
            title = str(item.get("title") or "").strip()
            if not title:
                continue
            distribution[title] = distribution.get(title, 0) + 1
    return dict(sorted(distribution.items(), key=lambda item: (-item[1], item[0])))


def build_summary(records: list[dict], expected_total: int) -> dict:
    total = expected_total or len(records)
    success = sum(1 for item in records if status_label(item) == "抽奖成功")
    risk = sum(1 for item in records if status_label(item) == "抽奖风控")
    failed = sum(1 for item in records if status_label(item) == "抽奖失败")
    abnormal = sum(1 for item in records if status_label(item) == "抽奖异常")
    total_lottery_results = sum(lottery_count(item) for item in records)
    prize_accounts = sum(1 for item in records if lottery_count(item) > 0)
    distribution = prize_distribution(records)
    success_rate = (success / total * 100) if total > 0 else 0.0
    return {
        "total": total,
        "success": success,
        "risk": risk,
        "failed": failed,
        "abnormal": abnormal,
        "problem_count": risk + failed + abnormal,
        "total_lottery_results": total_lottery_results,
        "prize_accounts": prize_accounts,
        "prize_distribution": distribution,
        "success_rate": success_rate,
    }


def build_stats_lines(summary: dict) -> list[str]:
    lines = [
        "📈 总体统计",
        f"  ├── 总账号数: {summary['total']}",
        f"  ├── 抽奖成功: {summary['success']}/{summary['total']}",
        f"  ├── 有中奖记录账号: {summary['prize_accounts']}",
        f"  ├── 抽奖结果总数: {summary['total_lottery_results']}",
        f"  └── 抽奖成功率: {format_percent(summary['success_rate'])}%",
    ]
    distribution = summary.get("prize_distribution") or {}
    if distribution:
        lines.append("🎁 奖品分布")
        for title, count in distribution.items():
            lines.append(f"  ├── {title}: {count}")
    return lines


def build_message(records: list[dict], manifest: dict, expected_total: int) -> tuple[str, dict]:
    sorted_records = sort_records(records)
    summary = build_summary(sorted_records, expected_total)
    problem_records = [record for record in sorted_records if is_problem_record(record)]

    if problem_records:
        lines = ["NO❗今天出现问题了捏"]
        for record in problem_records:
            lines.append(f"{record['username']}：{detail_reason(record)}❌")
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary

    if not sorted_records:
        lines = ["NO❗今天出现问题了捏", "未读取到任何抽奖结果❌"]
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary

    lines = ["喵喵~今天一切正常捏"]
    lines.extend(build_stats_lines(summary))
    return "\n".join(lines), summary


def fill_for_lottery_count(count: int):
    if count >= 3:
        return PatternFill("solid", fgColor="D9EAD3")
    if count == 2:
        return PatternFill("solid", fgColor="E2F0D9")
    if count == 1:
        return PatternFill("solid", fgColor="FFF2CC")
    return None


def font_for_status(label: str) -> Font:
    if label in {"抽奖失败", "抽奖异常", "抽奖风控"}:
        return Font(color="FFFFFF", bold=True)
    if label == "抽奖成功":
        return FONT_GREEN
    return FONT_DARK


def fill_for_status(label: str):
    if label in {"抽奖失败", "抽奖异常", "抽奖风控"}:
        return STATUS_RED_FILL
    return None


def is_bean_prize(title: str) -> bool:
    return "金豆" in str(title or "")


def is_normal_six_bean_prize(title: str) -> bool:
    compacted = re.sub(r"\s+", "", str(title or ""))
    return bool(re.fullmatch(r"6(?:个)?金豆", compacted))


def bean_expiry_date() -> str:
    return f"{datetime.now().year + 1}-12-31"


def lottery_columns(record: dict) -> list[str]:
    activity = normalize_activity_records(record.get("activity_records"))
    values = []
    rows = activity.get("lottery") or []
    for index in range(3):
        item = rows[index] if index < len(rows) else {}
        if item:
            title = str(item.get("title") or "").strip()
            expiry_date = str(item.get("expiry_date") or "").strip()
            if is_bean_prize(title):
                expiry_date = bean_expiry_date()
            values.extend([
                title,
                expiry_date,
            ])
        else:
            values.extend(["", ""])
    return values


def write_xlsx(path: str, records: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "抽奖汇总"
    headers = [
        "序号",
        "中奖数量",
        "账户",
        "组别",
        "抽奖情况",
        "抽奖IP",
        "抽奖1",
        "过期时间",
        "抽奖2",
        "过期时间",
        "抽奖3",
        "过期时间",
        "消费金额",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for index, record in enumerate(sort_records(records), start=1):
        label = status_label(record)
        row = [
            index,
            lottery_count(record),
            str(record.get("username") or ""),
            str(record.get("group_position") or ""),
            label,
            str(record.get("sign_ip") or ""),
        ] + lottery_columns(record) + [
            record_invoice_money(record),
        ]
        sheet.append(row)
        row_index = sheet.max_row
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        sheet.cell(row_index, 1).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row_index, 2).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row_index, 4).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row_index, 5).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row_index, 6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_index in range(7, 14):
            sheet.cell(row_index, column_index).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row_index, 2).number_format = "0"
        sheet.cell(row_index, 13).number_format = "0.00"
        fill = fill_for_lottery_count(lottery_count(record))
        if fill:
            sheet.cell(row_index, 2).fill = fill
        sheet.cell(row_index, 5).font = font_for_status(label)
        for column_index in (8, 10, 12):
            sheet.cell(row_index, column_index).font = FONT_RED if sheet.cell(row_index, column_index).value else FONT_DARK
        for prize_column in (7, 9, 11):
            prize_cell = sheet.cell(row_index, prize_column)
            expiry_cell = sheet.cell(row_index, prize_column + 1)
            title = str(prize_cell.value or "").strip()
            if not title:
                continue
            if not is_normal_six_bean_prize(title):
                prize_cell.fill = STATUS_RED_FILL
            if is_bean_prize(title):
                expiry_cell.value = bean_expiry_date()
                expiry_cell.font = FONT_GREEN

    sheet.freeze_panes = "A2"
    widths = {
        "A": 8,
        "B": 14,
        "C": 24,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 28,
        "H": 18,
        "I": 28,
        "J": 18,
        "K": 28,
        "L": 18,
        "M": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    workbook.save(path)


def split_text(text: str, limit: int = 3900) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts = []
    current = ""
    for line in text.splitlines(True):
        if len(current) + len(line) > limit and current:
            parts.append(current)
            current = ""
        current += line
    if current:
        parts.append(current)
    return parts


def telegram_credentials() -> tuple[str, str]:
    token = (
        os.getenv("TELEGRAM_BOT_TOKEN")
        or os.getenv("TG_BOT_TOKEN")
        or os.getenv("TELEGRAM_TOKEN")
        or os.getenv("TG_TOKEN")
        or ""
    ).strip()
    chat_id = (
        os.getenv("TELEGRAM_CHAT_ID")
        or os.getenv("TG_CHAT_ID")
        or os.getenv("TELEGRAM_TO")
        or os.getenv("TG_TO")
        or ""
    ).strip()
    return token, chat_id


def send_telegram_message(text: str) -> bool:
    token, chat_id = telegram_credentials()
    if not token or not chat_id:
        print("[telegram] skipped: Telegram bot token or chat id is empty")
        return False
    ok = True
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    for part in split_text(text):
        try:
            response = requests.post(url, json={"chat_id": chat_id, "text": part}, timeout=20)
            if response.status_code != 200:
                print(f"[telegram] sendMessage failed: HTTP {response.status_code} {response.text[:500]}")
                ok = False
        except Exception as exc:
            print(f"[telegram] sendMessage exception: {type(exc).__name__}: {exc}")
            ok = False
    return ok


def send_telegram_document(path: str) -> bool:
    token, chat_id = telegram_credentials()
    if not token or not chat_id:
        print("[telegram] skipped document: Telegram bot token or chat id is empty")
        return False
    if not os.path.exists(path):
        print(f"[telegram] skipped document: file not found: {path}")
        return False
    try:
        with open(path, "rb") as file:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": (os.path.basename(path), file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=40,
            )
        if response.status_code != 200:
            print(f"[telegram] sendDocument failed: HTTP {response.status_code} {response.text[:500]}")
            return False
        return True
    except Exception as exc:
        print(f"[telegram] sendDocument exception: {type(exc).__name__}: {exc}")
        return False


def send_email(subject: str, text: str) -> bool:
    host = os.getenv("SMTP_HOST")
    port = safe_int(os.getenv("SMTP_PORT", "465"), 465)
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASS")
    to_addr = os.getenv("SMTP_TO")
    from_addr = os.getenv("SMTP_FROM") or user
    if not host or not user or not password or not to_addr or not from_addr:
        return False
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = from_addr
    msg["To"] = to_addr
    try:
        if truthy(os.getenv("SMTP_USE_SSL", os.getenv("SMTP_SSL", "true"))):
            server = smtplib.SMTP_SSL(host, port, timeout=20)
        else:
            server = smtplib.SMTP(host, port, timeout=20)
            server.starttls(context=ssl.create_default_context())
        server.login(user, password)
        server.sendmail(from_addr, [to_addr], msg.as_string())
        server.quit()
        return True
    except Exception:
        return False


def parse_channels() -> list[str]:
    raw = (os.getenv("NOTIFY_CHANNELS") or "").strip()
    if raw:
        return [item.strip().lower() for item in raw.split(",") if item.strip()]
    channels = []
    token, chat_id = telegram_credentials()
    if token and chat_id:
        channels.append("telegram")
    if os.getenv("SMTP_HOST") and os.getenv("SMTP_TO"):
        channels.append("email")
    return channels


def is_enabled(env_name: str, default: str = "true") -> bool:
    return truthy(os.getenv(env_name, default))


def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else "results"
    account_lookup, expected_total = load_account_lookup()
    manifest = load_manifest(results_dir)
    output_xlsx = resolve_output_xlsx_path(results_dir, manifest)
    raw_records = load_results(results_dir, account_lookup)
    records = merge_records_with_expected(raw_records, account_lookup)
    message, summary = build_message(records, manifest, expected_total)

    channels = parse_channels()
    print(f"[notify] channels={','.join(channels) if channels else 'none'}")
    if "telegram" in channels and not all(telegram_credentials()):
        print("[notify] telegram requested but bot token/chat id is missing")
    send_tg_text = is_enabled("TELEGRAM_SEND_TEXT", "true")
    send_tg_xlsx = is_enabled("TELEGRAM_SEND_XLSX", "true")
    generate_xlsx = send_tg_xlsx or is_enabled("GENERATE_XLSX", "false")

    if generate_xlsx:
        write_xlsx(output_xlsx, records)

    sent = False
    if "telegram" in channels:
        if send_tg_text:
            sent = send_telegram_message(message) or sent
        if send_tg_xlsx:
            if not os.path.exists(output_xlsx):
                write_xlsx(output_xlsx, records)
            sent = send_telegram_document(output_xlsx) or sent
    if "email" in channels or "smtp" in channels:
        subject = f"{target_date_text(manifest)} 抽奖汇总"
        sent = send_email(subject, message) or sent

    print(message)
    print(
        f"[summary] total={summary['total']} success={summary['success']} "
        f"sent={'yes' if sent else 'no'} tg_text={'on' if send_tg_text else 'off'} "
        f"tg_xlsx={'on' if send_tg_xlsx else 'off'} xlsx_generated={'yes' if os.path.exists(output_xlsx) else 'no'}"
    )

    if truthy(os.getenv("FAIL_ON_FAILURE", "false")) and summary["problem_count"] > 0:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()

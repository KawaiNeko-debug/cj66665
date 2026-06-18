from __future__ import annotations

import glob
import json
import os
import smtplib
import ssl
import sys
from datetime import datetime
from email.header import Header
from email.mime.text import MIMEText
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(ROOT_DIR, ".env"))
LOCAL_TZ = ZoneInfo("Asia/Shanghai")

for stream_name in ("stdout", "stderr"):
    stream = getattr(sys, stream_name, None)
    if hasattr(stream, "reconfigure"):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


def truthy(value) -> bool:
    if value is True:
        return True
    if value in (False, None):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def safe_int(value, default=0) -> int:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def safe_float(value, default=0.0) -> float:
    try:
        return float(str(value).strip())
    except Exception:
        return default


def cell_text(value, limit: int = 32000) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    else:
        text = str(value)
    text = "\n".join(part.strip() for part in text.splitlines())
    if len(text) <= limit:
        return text
    return text[:limit] + f"...(truncated,len={len(text)})"


def parse_account_line(line: str):
    if "," in line:
        username, _ = line.split(",", 1)
    elif "----" in line:
        username, _ = line.split("----", 1)
    else:
        return None
    username = username.strip()
    return username or None


def expected_accounts() -> tuple[dict[tuple[int, int], str], int]:
    lookup = {}
    raw = os.getenv("ACCOUNTS_BATCH1", "") or os.getenv("ACCOUNTS", "") or ""
    for account_index, line in enumerate(raw.splitlines(), start=1):
        username = parse_account_line(line.strip())
        if username:
            lookup[(1, account_index)] = username
    return lookup, len(lookup)


def find_json_files(results_dir: str) -> list[str]:
    paths = []
    for path in glob.glob(os.path.join(results_dir, "**", "*.json"), recursive=True):
        if os.path.basename(path).lower() == "manifest.json":
            continue
        if os.path.isfile(path):
            paths.append(path)
    return sorted(paths)


def onebean_records(row: dict) -> list[dict]:
    activity = row.get("activity_records") if isinstance(row.get("activity_records"), dict) else {}
    records = activity.get("onebean")
    if isinstance(records, list):
        return [item for item in records if isinstance(item, dict)]
    return []


def prize_quantity(records: list[dict]) -> int:
    return sum(max(1, safe_int(item.get("quantity"), 1)) for item in records)


def normalize_record(row: dict, payload: dict, account_lookup: dict[tuple[int, int], str]) -> dict:
    group_number = safe_int(row.get("group_number") or payload.get("group_number"), 1)
    account_index = safe_int(row.get("account_index"), 0)
    records = onebean_records(row)
    username = (
        row.get("username")
        or row.get("masked_username")
        or account_lookup.get((group_number, account_index))
        or f"账号{account_index}"
    )
    success = truthy(row.get("sign_success")) and bool(records)
    return {
        "group_number": group_number,
        "account_index": account_index,
        "group_name": row.get("group_name") or payload.get("group_name") or "onebean",
        "group_position": row.get("group_position") or f"{group_number}组账号{account_index}",
        "username": str(username),
        "success": success,
        "status": row.get("sign_status") or ("领取成功" if success else "领取失败"),
        "reason": row.get("detail_reason") or row.get("sign_status") or "",
        "password_error": truthy(row.get("password_error")),
        "risk_controlled": truthy(row.get("risk_controlled")),
        "sign_time": row.get("sign_time") or "",
        "records": records,
    }


def load_results(results_dir: str, account_lookup: dict[tuple[int, int], str]) -> list[dict]:
    records = []
    for path in find_json_files(results_dir):
        try:
            with open(path, "r", encoding="utf-8") as file:
                payload = json.load(file)
        except Exception:
            continue
        rows = payload.get("results") if isinstance(payload, dict) else None
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, dict):
                records.append(normalize_record(row, payload, account_lookup))
    return records


def merge_expected(records: list[dict], account_lookup: dict[tuple[int, int], str]) -> list[dict]:
    indexed = {(item["group_number"], item["account_index"]): item for item in records if item.get("account_index")}
    if not account_lookup:
        return list(indexed.values())
    merged = []
    for key in sorted(account_lookup):
        item = indexed.pop(key, None)
        if item is None:
            merged.append(
                {
                    "group_number": key[0],
                    "account_index": key[1],
                    "group_name": "onebean",
                    "group_position": f"{key[0]}组账号{key[1]}",
                    "username": account_lookup[key],
                    "success": False,
                    "status": "未回传结果",
                    "reason": "汇总时未找到该账号 result.json",
                    "password_error": False,
                    "risk_controlled": False,
                    "sign_time": "",
                    "records": [],
                }
            )
        else:
            merged.append(item)
    merged.extend(indexed.values())
    return merged


def prize_distribution(records: list[dict]) -> dict[str, int]:
    dist: dict[str, int] = {}
    for row in records:
        for item in row.get("records") or []:
            title = str(item.get("title") or item.get("prize_name") or "未知奖品").strip()
            if not title:
                continue
            dist[title] = dist.get(title, 0) + max(1, safe_int(item.get("quantity"), 1))
    return dict(sorted(dist.items(), key=lambda item: (-item[1], item[0])))


def build_summary(records: list[dict], expected_total: int) -> dict:
    total = expected_total or len(records)
    success = sum(1 for row in records if row.get("success"))
    failed = max(0, total - success)
    total_prizes = sum(prize_quantity(row.get("records") or []) for row in records)
    return {
        "total": total,
        "success": success,
        "failed": failed,
        "total_prizes": total_prizes,
        "success_rate": (success / total * 100) if total else 0.0,
        "distribution": prize_distribution(records),
    }


def format_percent(value: float) -> str:
    return f"{value:.1f}".rstrip("0").rstrip(".")


def build_stats_lines(summary: dict) -> list[str]:
    lines = [
        "📈 总体统计",
        f"  ├── 总账号数: {summary['total']}",
        f"  ├── 领取成功: {summary['success']}/{summary['total']}",
        f"  ├── 失败账号: {summary['failed']}",
        f"  ├── 奖品总数: {summary['total_prizes']}",
        f"  └── 成功率: {format_percent(summary['success_rate'])}%",
    ]
    distribution = summary.get("distribution") or {}
    if distribution:
        lines.append("🎁 奖品分布")
        for title, count in distribution.items():
            lines.append(f"  ├── {title}: {count}")
    return lines


def build_message(records: list[dict], expected_total: int) -> tuple[str, dict]:
    summary = build_summary(records, expected_total)
    problems = [row for row in records if not row.get("success")]
    if problems:
        lines = ["NO❌今天出现问题了捏"]
        for row in problems:
            lines.append(f"{row.get('username')}: {row.get('reason') or row.get('status') or '领取失败'}❌")
        lines.extend(build_stats_lines(summary))
        return "\n".join(lines), summary
    lines = ["喵喵~今天一切正常捏"]
    lines.extend(build_stats_lines(summary))
    return "\n".join(lines), summary


def joined(records: list[dict], key: str) -> str:
    values = []
    for item in records:
        value = item.get(key)
        if value not in (None, ""):
            values.append(str(value))
    return "\n".join(values)


def write_xlsx(path: str, records: list[dict]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1豆汇总"
    headers = [
        "序号",
        "账号",
        "组别",
        "领取状态",
        "详情",
        "奖品数量",
        "奖品ID",
        "奖品名称",
        "面额",
        "最低消费",
        "开始日期",
        "过期时间",
        "业务线",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    success_fill = PatternFill("solid", fgColor="D9EAD3")
    fail_fill = PatternFill("solid", fgColor="F8696B")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    sorted_records = sorted(records, key=lambda item: (safe_int(item.get("group_number"), 999), safe_int(item.get("account_index"), 999)))
    for index, row in enumerate(sorted_records, start=1):
        prize_rows = row.get("records") or []
        values = [
            index,
            row.get("username") or "",
            row.get("group_position") or "",
            "领取成功" if row.get("success") else row.get("status") or "领取失败",
            row.get("reason") or "",
            prize_quantity(prize_rows),
            joined(prize_rows, "customer_coupon_id") or joined(prize_rows, "coupon_id"),
            joined(prize_rows, "title"),
            joined(prize_rows, "denomination"),
            joined(prize_rows, "min_consume_money"),
            joined(prize_rows, "start_date"),
            joined(prize_rows, "expiry_date"),
            joined(prize_rows, "business_line"),
        ]
        sheet.append(values)
        row_index = sheet.max_row
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row_index, 4).fill = success_fill if row.get("success") else fail_fill
        if not row.get("success"):
            sheet.cell(row_index, 4).font = Font(color="FFFFFF", bold=True)
        sheet.cell(row_index, 6).number_format = "0"

    widths = {
        "A": 8,
        "B": 24,
        "C": 16,
        "D": 14,
        "E": 34,
        "F": 12,
        "G": 28,
        "H": 42,
        "I": 12,
        "J": 12,
        "K": 14,
        "L": 14,
        "M": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    workbook.save(path)


def split_text(text: str, limit: int = 3900) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts = []
    current = ""
    for line in text.splitlines(True):
        if current and len(current) + len(line) > limit:
            parts.append(current)
            current = ""
        current += line
    if current:
        parts.append(current)
    return parts


def telegram_credentials() -> tuple[str, str]:
    token = (
        os.getenv("TELEGRAM_BOT_TOKEN")
        or os.getenv("TG_BOT_TOKEN")
        or os.getenv("TELEGRAM_TOKEN")
        or os.getenv("TG_TOKEN")
        or ""
    ).strip()
    chat_id = (
        os.getenv("TELEGRAM_CHAT_ID")
        or os.getenv("TG_CHAT_ID")
        or os.getenv("TELEGRAM_TO")
        or os.getenv("TG_TO")
        or ""
    ).strip()
    return token, chat_id


def send_telegram_message(text: str) -> bool:
    token, chat_id = telegram_credentials()
    if not token or not chat_id:
        print("[telegram] skipped: Telegram bot token or chat id is empty")
        return False
    ok = True
    for part in split_text(text):
        try:
            response = requests.post(f"https://api.telegram.org/bot{token}/sendMessage", json={"chat_id": chat_id, "text": part}, timeout=20)
            if response.status_code != 200:
                print(f"[telegram] sendMessage failed: HTTP {response.status_code} {response.text[:500]}")
                ok = False
        except Exception as exc:
            print(f"[telegram] sendMessage exception: {type(exc).__name__}: {exc}")
            ok = False
    return ok


def send_telegram_document(path: str) -> bool:
    token, chat_id = telegram_credentials()
    if not token or not chat_id:
        print("[telegram] skipped document: Telegram bot token or chat id is empty")
        return False
    if not os.path.exists(path):
        print(f"[telegram] skipped document: file not found: {path}")
        return False
    try:
        with open(path, "rb") as file:
            response = requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": (os.path.basename(path), file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=40,
            )
        if response.status_code != 200:
            print(f"[telegram] sendDocument failed: HTTP {response.status_code} {response.text[:500]}")
            return False
        return True
    except Exception as exc:
        print(f"[telegram] sendDocument exception: {type(exc).__name__}: {exc}")
        return False


def send_email(subject: str, text: str) -> bool:
    host = os.getenv("SMTP_HOST")
    user = os.getenv("SMTP_USER")
    password = os.getenv("SMTP_PASS")
    to_addr = os.getenv("SMTP_TO")
    from_addr = os.getenv("SMTP_FROM") or user
    if not host or not user or not password or not to_addr or not from_addr:
        return False
    msg = MIMEText(text, "plain", "utf-8")
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = from_addr
    msg["To"] = to_addr
    try:
        if truthy(os.getenv("SMTP_USE_SSL", os.getenv("SMTP_SSL", "true"))):
            server = smtplib.SMTP_SSL(host, safe_int(os.getenv("SMTP_PORT"), 465), timeout=20)
        else:
            server = smtplib.SMTP(host, safe_int(os.getenv("SMTP_PORT"), 587), timeout=20)
            server.starttls(context=ssl.create_default_context())
        server.login(user, password)
        server.sendmail(from_addr, [to_addr], msg.as_string())
        server.quit()
        return True
    except Exception:
        return False


def output_xlsx_path(results_dir: str) -> str:
    configured = (os.getenv("OUTPUT_XLSX_PATH") or "").strip()
    if configured:
        return configured
    return os.path.join(results_dir, "1豆-summary.xlsx")


def main():
    results_dir = sys.argv[1] if len(sys.argv) > 1 else "merged"
    account_lookup, expected_total = expected_accounts()
    records = merge_expected(load_results(results_dir, account_lookup), account_lookup)
    message, summary = build_message(records, expected_total)
    output_xlsx = output_xlsx_path(results_dir)
    send_tg_text = truthy(os.getenv("TELEGRAM_SEND_TEXT", "true"))
    send_tg_xlsx = truthy(os.getenv("TELEGRAM_SEND_XLSX", "true"))
    generate_xlsx = send_tg_xlsx or truthy(os.getenv("GENERATE_XLSX", "false"))
    if generate_xlsx:
        write_xlsx(output_xlsx, records)
    sent = False
    channels = [item.strip().lower() for item in (os.getenv("NOTIFY_CHANNELS") or "telegram").split(",") if item.strip()]
    print(f"[notify] channels={','.join(channels) if channels else 'none'}")
    if "telegram" in channels:
        if send_tg_text:
            sent = send_telegram_message(message) or sent
        if send_tg_xlsx:
            if not os.path.exists(output_xlsx):
                write_xlsx(output_xlsx, records)
            sent = send_telegram_document(output_xlsx) or sent
    if "email" in channels or "smtp" in channels:
        sent = send_email(f"{datetime.now(LOCAL_TZ).strftime('%Y-%m-%d')} 1豆汇总", message) or sent
    print(message)
    print(
        f"[summary] total={summary['total']} success={summary['success']} "
        f"sent={'yes' if sent else 'no'} tg_text={'on' if send_tg_text else 'off'} "
        f"tg_xlsx={'on' if send_tg_xlsx else 'off'} xlsx_generated={'yes' if os.path.exists(output_xlsx) else 'no'}"
    )
    if truthy(os.getenv("FAIL_ON_FAILURE", "false")) and summary["failed"] > 0:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
